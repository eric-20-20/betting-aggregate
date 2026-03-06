#!/usr/bin/env python3
"""Export analytics reports to a multi-sheet Excel workbook."""

import json
import math
import os
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

OCCURRENCES_PATH = Path("data/analysis/graded_occurrences_latest.jsonl")
SIGNALS_PATH = Path("data/ledger/signals_latest.jsonl")
GRADES_PATH = Path("data/ledger/grades_latest.jsonl")

REPORTS_DIR = Path("data/reports")
TRENDS_DIR = REPORTS_DIR / "trends"
OUTPUT_DIR = Path("data/analytics")
OUTPUT_FILE = OUTPUT_DIR / "analytics_report.xlsx"

# ── Sheet definitions ───────────────────────────────────────────────
# Each entry: (sheet_name, json_path, rows_key, columns)
# columns: list of (header, json_key, format)
#   format: None=general, "pct"=percentage, "units"=+/- decimal, "int"=integer

PCT = "pct"
UNITS = "units"
INT = "int"

SHEETS = [
    (
        "Source Records",
        REPORTS_DIR / "by_source_record.json",
        "rows",
        [
            ("Source", "source_id", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
            ("Avg Odds", "avg_odds", UNITS),
        ],
    ),
    (
        "Source x Market",
        REPORTS_DIR / "by_source_record_market.json",
        "rows",
        [
            ("Source", "source_id", None),
            ("Market", "market_type", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
        ],
    ),
    (
        "Source Combos",
        REPORTS_DIR / "by_sources_combo_record.json",
        "rows",
        [
            ("Sources Combo", "sources_combo", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
        ],
    ),
    (
        "Combo x Market",
        REPORTS_DIR / "by_sources_combo_market.json",
        "rows",
        [
            ("Sources Combo", "sources_combo", None),
            ("Market", "market_type", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
        ],
    ),
    (
        "Expert Records",
        REPORTS_DIR / "by_expert_record.json",
        "rows",
        [
            ("Expert", "expert", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
        ],
    ),
    (
        "Stat Types",
        REPORTS_DIR / "by_stat_type.json",
        "rows",
        [
            ("Stat Type", "stat_type", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
        ],
    ),
    (
        "Stat x Source",
        REPORTS_DIR / "by_stat_type_source.json",
        "rows",
        [
            ("Source", "source_id", None),
            ("Stat Type", "stat_type", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
        ],
    ),
    (
        "Line Buckets",
        REPORTS_DIR / "by_line_bucket.json",
        "rows",
        [
            ("Line Bucket", "line_bucket", None),
            ("Market", "market_type", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
        ],
    ),
    (
        "Score Buckets",
        REPORTS_DIR / "by_score_bucket.json",
        "rows",
        [
            ("Score Bucket", "score_bucket", None),
            ("Market", "market_type", None),
            ("N Graded", "n_graded", INT),
            ("N Total", "n_total", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
        ],
    ),
    (
        "Consensus Strength",
        TRENDS_DIR / "consensus_strength.json",
        "rows",
        [
            ("Consensus", "consensus_strength", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("Wilson Lower", "wilson_lower", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
        ],
    ),
    (
        "Market Type",
        TRENDS_DIR / "market_type.json",
        "by_market",
        [
            ("Market", "market_type", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("Wilson Lower", "wilson_lower", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
        ],
    ),
    (
        "Market x Consensus",
        TRENDS_DIR / "market_type.json",
        "by_market_x_consensus",
        [
            ("Market", "market_type", None),
            ("Consensus", "consensus_strength", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("Wilson Lower", "wilson_lower", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
            ("Sample Flag", "sample_flag", None),
        ],
    ),
    (
        "By Month",
        TRENDS_DIR / "by_month.json",
        "rows",
        [
            ("Month", "month", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("Wilson Lower", "wilson_lower", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
            ("Sample Flag", "sample_flag", None),
        ],
    ),
    (
        "By Day of Week",
        TRENDS_DIR / "by_day_of_week.json",
        "rows",
        [
            ("Day of Week", "day_of_week", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("Wilson Lower", "wilson_lower", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
            ("Sample Flag", "sample_flag", None),
        ],
    ),
    (
        "Source Surface",
        TRENDS_DIR / "by_source_surface.json",
        "rows",
        [
            ("Source Surface", "source_surface", None),
            ("N", "n", INT),
            ("Wins", "wins", INT),
            ("Losses", "losses", INT),
            ("Pushes", "pushes", INT),
            ("Win %", "win_pct", PCT),
            ("Wilson Lower", "wilson_lower", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Bets w/ Units", "bets_with_units", INT),
            ("Sample Flag", "sample_flag", None),
        ],
    ),
    (
        "Top Trends",
        TRENDS_DIR / "top_trends_summary.json",
        "trends",
        [
            ("Description", "description", None),
            ("Report", "report", None),
            ("N", "n", INT),
            ("Win %", "win_pct", PCT),
            ("Wilson Lower", "wilson_lower", PCT),
            ("ROI", "roi", PCT),
            ("Net Units", "net_units", UNITS),
            ("Edge over 50%", "edge_over_50", PCT),
            ("Sample Flag", "sample_flag", None),
        ],
    ),
]


def auto_size_columns(ws):
    """Auto-size columns based on content width."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def write_sheet(wb, sheet_name, json_path, rows_key, columns):
    """Write a single sheet from a JSON report."""
    if not json_path.exists():
        print(f"  SKIP {sheet_name}: {json_path} not found")
        return

    with open(json_path) as f:
        data = json.load(f)

    rows = data.get(rows_key, [])
    if not rows:
        print(f"  SKIP {sheet_name}: no rows in '{rows_key}'")
        return

    # Sort by sample size descending
    size_key = "n" if "n" in rows[0] else "n_graded" if "n_graded" in rows[0] else None
    if size_key:
        rows = sorted(rows, key=lambda r: r.get(size_key, 0), reverse=True)

    ws = wb.create_sheet(title=sheet_name)
    bold = Font(bold=True)

    # Header row
    for col_idx, (header, _, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, (_, key, fmt) in enumerate(columns, 1):
            val = row_data.get(key)
            cell = ws.cell(row=row_idx, column=col_idx)

            if val is None:
                cell.value = ""
                continue

            cell.value = val

            if fmt == PCT and isinstance(val, (int, float)):
                cell.number_format = '0.00%'
                # Convert decimal to percentage if needed (values like 0.51 → 51%)
                # openpyxl percentage format multiplies by 100, so keep raw decimal
            elif fmt == UNITS and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            elif fmt == INT and isinstance(val, (int, float)):
                cell.number_format = '#,##0'

    auto_size_columns(ws)
    print(f"  OK   {sheet_name}: {len(rows)} rows")


def wilson_lower(wins, n, z=1.96):
    if n == 0:
        return 0.0
    p = wins / n
    denom = 1 + z * z / n
    center = p + z * z / (2 * n)
    spread = z * math.sqrt((p * (1 - p) + z * z / (4 * n)) / n)
    return (center - spread) / denom


def parse_month(dk):
    parts = (dk or "").split(":")
    return f"{parts[1]}-{parts[2]}" if len(parts) >= 3 else "unknown"


def norm_direction(d):
    if not d:
        return ""
    u = str(d).upper()
    return u if u in ("OVER", "UNDER") else ""


def build_combo_direction_sheet(wb):
    """Build Combo x Market x Direction sheet directly from graded_occurrences_latest.jsonl."""
    if not OCCURRENCES_PATH.exists():
        print(f"  SKIP Combo x Mkt x Dir: {OCCURRENCES_PATH} not found")
        return

    # Accumulate stats
    groups = defaultdict(lambda: {"w": 0, "l": 0, "months": defaultdict(lambda: [0, 0])})

    with open(OCCURRENCES_PATH) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            result = row.get("result")
            if result not in ("WIN", "LOSS"):
                continue
            combo = row.get("sources_combo") or ""
            mkt = row.get("market_type") or ""
            if mkt not in ("player_prop", "spread", "total", "moneyline"):
                continue
            direction = norm_direction(row.get("direction"))
            month = parse_month(row.get("day_key", ""))
            key = (combo, mkt, direction)
            if result == "WIN":
                groups[key]["w"] += 1
                groups[key]["months"][month][0] += 1
            else:
                groups[key]["l"] += 1
                groups[key]["months"][month][1] += 1

    # Build rows
    rows = []
    for (combo, mkt, direction), g in groups.items():
        w, l = g["w"], g["l"]
        n = w + l
        if n < 5:
            continue
        win_pct = w / n
        wl = wilson_lower(w, n)

        # Monthly breakdown string
        month_parts = []
        for month in sorted(g["months"].keys()):
            mw, ml = g["months"][month]
            mn = mw + ml
            if mn >= 5:
                month_parts.append(f"{month}:{mw}-{ml}({int(mw/mn*100)}%)")
        monthly_str = "  ".join(month_parts)

        # Consistency flag
        month_pcts = [mw / (mw + ml) for mw, ml in g["months"].values() if (mw + ml) >= 5]
        if len(month_pcts) == 1:
            flag = "single_month"
        elif len(month_pcts) >= 3 and sum(1 for p in month_pcts if p >= 0.55) <= 1:
            flag = "inconsistent"
        else:
            flag = ""

        n_months = len(month_pcts)

        rows.append({
            "sources_combo": combo,
            "market_type": mkt,
            "direction": direction or "(all)",
            "n": n,
            "wins": w,
            "losses": l,
            "win_pct": win_pct,
            "wilson_lower": wl,
            "n_months": n_months,
            "consistency": flag,
            "monthly_detail": monthly_str,
        })

    # Sort by wilson_lower desc
    rows.sort(key=lambda r: r["wilson_lower"], reverse=True)

    ws = wb.create_sheet(title="Combo x Mkt x Dir")
    bold = Font(bold=True)

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    columns = [
        ("Sources Combo", "sources_combo", None),
        ("Market", "market_type", None),
        ("Direction", "direction", None),
        ("N", "n", INT),
        ("Wins", "wins", INT),
        ("Losses", "losses", INT),
        ("Win %", "win_pct", PCT),
        ("Wilson Lower", "wilson_lower", PCT),
        ("Months", "n_months", INT),
        ("Consistency", "consistency", None),
        ("Monthly Detail", "monthly_detail", None),
    ]

    for col_idx, (header, _, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, (_, key, fmt) in enumerate(columns, 1):
            val = row_data.get(key)
            cell = ws.cell(row=row_idx, column=col_idx)
            if val is None:
                cell.value = ""
                continue
            cell.value = val
            if fmt == PCT and isinstance(val, (int, float)):
                cell.number_format = "0.00%"
            elif fmt == INT and isinstance(val, (int, float)):
                cell.number_format = "#,##0"

        # Color rows by win_pct
        wp = row_data.get("win_pct", 0)
        n = row_data.get("n", 0)
        if n >= 20:
            if wp >= 0.58:
                fill = green_fill
            elif wp <= 0.44:
                fill = red_fill
            elif wp >= 0.53:
                fill = yellow_fill
            else:
                fill = None
            if fill:
                for col_idx in range(1, len(columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    auto_size_columns(ws)
    # Make monthly detail column wider
    ws.column_dimensions[get_column_letter(len(columns))].width = 80
    print(f"  OK   Combo x Mkt x Dir: {len(rows)} rows (built from occurrences)")


def build_monthly_trends_sheet(wb):
    """Build per-combo monthly win rate table — rows=combo×mkt×dir, cols=months."""
    if not OCCURRENCES_PATH.exists():
        return

    groups = defaultdict(lambda: defaultdict(lambda: [0, 0]))
    all_months = set()

    with open(OCCURRENCES_PATH) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            result = row.get("result")
            if result not in ("WIN", "LOSS"):
                continue
            combo = row.get("sources_combo") or ""
            mkt = row.get("market_type") or ""
            if mkt not in ("player_prop", "spread", "total", "moneyline"):
                continue
            direction = norm_direction(row.get("direction"))
            month = parse_month(row.get("day_key", ""))
            key = (combo, mkt, direction or "(all)")
            all_months.add(month)
            if result == "WIN":
                groups[key][month][0] += 1
            else:
                groups[key][month][1] += 1

    sorted_months = sorted(m for m in all_months if m != "unknown")

    # Only include combos with n>=20 total
    summary = []
    for key, month_data in groups.items():
        total_w = sum(v[0] for v in month_data.values())
        total_l = sum(v[1] for v in month_data.values())
        n = total_w + total_l
        if n < 20:
            continue
        summary.append((key, month_data, total_w, total_l, n))

    summary.sort(key=lambda x: wilson_lower(x[2], x[4]), reverse=True)

    ws = wb.create_sheet(title="Monthly Trends")
    bold = Font(bold=True)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")

    # Header
    headers = ["Sources Combo", "Market", "Direction", "Total N", "Total Win%", "Wilson"] + sorted_months
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold

    for row_idx, ((combo, mkt, direction), month_data, total_w, total_l, n) in enumerate(summary, 2):
        ws.cell(row=row_idx, column=1, value=combo)
        ws.cell(row=row_idx, column=2, value=mkt)
        ws.cell(row=row_idx, column=3, value=direction)
        ws.cell(row=row_idx, column=4, value=n).number_format = "#,##0"
        wp_cell = ws.cell(row=row_idx, column=5, value=total_w / n)
        wp_cell.number_format = "0.0%"
        wl_cell = ws.cell(row=row_idx, column=6, value=wilson_lower(total_w, n))
        wl_cell.number_format = "0.000"

        for col_idx, month in enumerate(sorted_months, 7):
            mw, ml = month_data.get(month, [0, 0])
            mn = mw + ml
            if mn == 0:
                ws.cell(row=row_idx, column=col_idx, value="")
                continue
            cell = ws.cell(row=row_idx, column=col_idx, value=mw / mn)
            cell.number_format = "0%"
            # Color individual month cells
            mp = mw / mn
            if mn >= 10:
                if mp >= 0.60:
                    cell.fill = green_fill
                elif mp <= 0.42:
                    cell.fill = red_fill
                elif mp >= 0.53:
                    cell.fill = yellow_fill

    auto_size_columns(ws)
    print(f"  OK   Monthly Trends: {len(summary)} combo×mkt×dir rows × {len(sorted_months)} months")


def _load_ledger_records():
    """Load and merge signals + grades ledger into flat records list."""
    if not SIGNALS_PATH.exists() or not GRADES_PATH.exists():
        return []
    signals = {}
    with open(SIGNALS_PATH) as f:
        for line in f:
            s = json.loads(line)
            signals[s["signal_id"]] = s
    records = []
    with open(GRADES_PATH) as f:
        for line in f:
            g = json.loads(line)
            result = g.get("result")
            if result not in ("WIN", "LOSS"):
                continue
            sid = g.get("signal_id")
            s = signals.get(sid)
            if not s:
                continue
            day_key = s.get("day_key", "")
            parts = day_key.split(":")
            month = f"{parts[1]}-{parts[2]}" if len(parts) >= 3 else "unknown"
            records.append({
                "month": month,
                "result": result,
                "market_type": s.get("market_type"),
                "direction": s.get("direction"),
                "sources_combo": s.get("sources_combo", ""),
                "sources_present": s.get("sources_present", []),
                "line": s.get("line"),
            })
    return records


def _is_team_abbrev(d):
    """Return True if direction looks like a 3-letter NBA team abbreviation."""
    return bool(d and len(d) == 3 and d.isalpha() and d.isupper())


def build_team_trends_sheet(wb, records):
    """Build Team Trends sheet: team × market × source combo with Wilson scores."""
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    bold = Font(bold=True)

    # Accumulate: key = (team, market, sources_combo)
    key_stats = defaultdict(lambda: [0, 0])
    key_monthly = defaultdict(lambda: defaultdict(lambda: [0, 0]))

    for r in records:
        mkt = r["market_type"]
        dire = r["direction"]
        if mkt not in ("spread", "moneyline"):
            continue
        if not _is_team_abbrev(dire):
            continue
        key = (dire, mkt, r["sources_combo"])
        win = 1 if r["result"] == "WIN" else 0
        key_stats[key][0] += win
        key_stats[key][1] += (1 - win)
        key_monthly[key][r["month"]][0] += win
        key_monthly[key][r["month"]][1] += (1 - win)

    rows = []
    for (team, mkt, combo), (w, l) in key_stats.items():
        n = w + l
        if n < 5:
            continue
        win_pct = w / n
        wl = wilson_lower(w, n)
        monthly = key_monthly[(team, mkt, combo)]
        month_parts = []
        month_pcts = []
        for month in sorted(monthly.keys()):
            mw, ml = monthly[month]
            mn = mw + ml
            if mn >= 5:
                month_pcts.append(mw / mn)
                month_parts.append(f"{month}:{mw}-{ml}({int(mw/mn*100)}%)")
        if len(month_pcts) == 1:
            consistency = "single_month"
        elif len(month_pcts) >= 3 and sum(1 for p in month_pcts if p >= 0.55) <= 1:
            consistency = "inconsistent"
        else:
            consistency = ""
        rows.append({
            "team": team,
            "market_type": mkt,
            "sources_combo": combo,
            "n": n,
            "wins": w,
            "losses": l,
            "win_pct": win_pct,
            "wilson_lower": wl,
            "n_months": len(month_pcts),
            "consistency": consistency,
            "monthly_detail": "  ".join(month_parts),
        })

    rows.sort(key=lambda r: r["wilson_lower"], reverse=True)

    ws = wb.create_sheet(title="Team Trends")
    columns = [
        ("Team", "team", None),
        ("Market", "market_type", None),
        ("Sources Combo", "sources_combo", None),
        ("N", "n", INT),
        ("Wins", "wins", INT),
        ("Losses", "losses", INT),
        ("Win %", "win_pct", PCT),
        ("Wilson Lower", "wilson_lower", PCT),
        ("Months", "n_months", INT),
        ("Consistency", "consistency", None),
        ("Monthly Detail", "monthly_detail", None),
    ]

    for col_idx, (header, _, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, (_, key, fmt) in enumerate(columns, 1):
            val = row_data.get(key)
            cell = ws.cell(row=row_idx, column=col_idx)
            if val is None:
                cell.value = ""
                continue
            cell.value = val
            if fmt == PCT and isinstance(val, (int, float)):
                cell.number_format = "0.00%"
            elif fmt == INT and isinstance(val, (int, float)):
                cell.number_format = "#,##0"

        wp = row_data.get("win_pct", 0)
        n = row_data.get("n", 0)
        if n >= 20:
            if wp >= 0.58:
                fill = green_fill
            elif wp <= 0.44:
                fill = red_fill
            elif wp >= 0.53:
                fill = yellow_fill
            else:
                fill = None
            if fill:
                for col_idx in range(1, len(columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    auto_size_columns(ws)
    ws.column_dimensions[get_column_letter(len(columns))].width = 80
    print(f"  OK   Team Trends: {len(rows)} rows")


def build_team_monthly_sheet(wb, records):
    """Build Team Monthly pivot: rows=team×market×combo, cols=months."""
    bold = Font(bold=True)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")

    groups = defaultdict(lambda: defaultdict(lambda: [0, 0]))
    all_months = set()

    for r in records:
        mkt = r["market_type"]
        dire = r["direction"]
        if mkt not in ("spread", "moneyline"):
            continue
        if not _is_team_abbrev(dire):
            continue
        key = (dire, mkt, r["sources_combo"])
        month = r["month"]
        all_months.add(month)
        win = 1 if r["result"] == "WIN" else 0
        groups[key][month][0] += win
        groups[key][month][1] += (1 - win)

    sorted_months = sorted(m for m in all_months if m != "unknown")

    summary = []
    for key, month_data in groups.items():
        total_w = sum(v[0] for v in month_data.values())
        total_l = sum(v[1] for v in month_data.values())
        n = total_w + total_l
        if n < 10:
            continue
        summary.append((key, month_data, total_w, total_l, n))

    summary.sort(key=lambda x: wilson_lower(x[2], x[4]), reverse=True)

    ws = wb.create_sheet(title="Team Monthly")
    headers = ["Team", "Market", "Sources Combo", "Total N", "Total Win%", "Wilson"] + sorted_months
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold

    for row_idx, ((team, mkt, combo), month_data, total_w, total_l, n) in enumerate(summary, 2):
        ws.cell(row=row_idx, column=1, value=team)
        ws.cell(row=row_idx, column=2, value=mkt)
        ws.cell(row=row_idx, column=3, value=combo)
        ws.cell(row=row_idx, column=4, value=n).number_format = "#,##0"
        wp_cell = ws.cell(row=row_idx, column=5, value=total_w / n)
        wp_cell.number_format = "0.0%"
        wl_cell = ws.cell(row=row_idx, column=6, value=wilson_lower(total_w, n))
        wl_cell.number_format = "0.000"

        for col_idx, month in enumerate(sorted_months, 7):
            mw, ml = month_data.get(month, [0, 0])
            mn = mw + ml
            if mn == 0:
                ws.cell(row=row_idx, column=col_idx, value="")
                continue
            cell = ws.cell(row=row_idx, column=col_idx, value=mw / mn)
            cell.number_format = "0%"
            mp = mw / mn
            if mn >= 8:
                if mp >= 0.60:
                    cell.fill = green_fill
                elif mp <= 0.42:
                    cell.fill = red_fill
                elif mp >= 0.53:
                    cell.fill = yellow_fill

    auto_size_columns(ws)
    print(f"  OK   Team Monthly: {len(summary)} team×market×combo rows × {len(sorted_months)} months")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    print("Building analytics Excel export...")
    for sheet_name, json_path, rows_key, columns in SHEETS:
        write_sheet(wb, sheet_name, json_path, rows_key, columns)

    # Add computed sheets from raw occurrences
    build_combo_direction_sheet(wb)
    build_monthly_trends_sheet(wb)

    # Add team trend sheets from ledger
    print("  Loading ledger records for team trend sheets...")
    ledger_records = _load_ledger_records()
    print(f"  Loaded {len(ledger_records)} graded WIN/LOSS records")
    build_team_trends_sheet(wb, ledger_records)
    build_team_monthly_sheet(wb, ledger_records)

    wb.save(str(OUTPUT_FILE))
    print(f"\nSaved: {OUTPUT_FILE}")
    print(f"Sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
