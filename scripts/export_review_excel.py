#!/usr/bin/env python3
"""Export a combined NBA + MLB review workbook for manual inspection.

Reads existing CSVs from data/reports/sports_review_pack/ and
data/reports/pattern_audit/, writes a single Excel workbook.

Read-only — does not modify pipeline data or scoring.

Usage:
    python3 scripts/export_review_excel.py
"""

from __future__ import annotations

import csv
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parents[1]
REVIEW_DIR = REPO_ROOT / "data" / "reports" / "sports_review_pack"
AUDIT_DIR = REPO_ROOT / "data" / "reports" / "pattern_audit"
OUTPUT_PATH = REVIEW_DIR / "sports_review_workbook.xlsx"

# Styles
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
CLEAN_FILL = PatternFill("solid", fgColor="C6EFCE")
BOLD = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)


def read_csv_rows(path: Path) -> tuple:
    """Return (headers, rows) from a CSV file."""
    if not path.exists():
        return [], []
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        rows = list(reader)
    return headers, rows


def add_data_sheet(wb: Workbook, name: str, headers: list, rows: list,
                   severity_col: Optional[int] = None, max_rows: int = 50000) -> int:
    """Add a sheet with data, headers, filters, and frozen top row."""
    ws = wb.create_sheet(title=name)

    if not headers:
        ws.append(["No data available"])
        return 0

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Write data (capped)
    capped = rows[:max_rows]
    for row_idx, row in enumerate(capped, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_coerce(value))

            # Color severity column
            if severity_col is not None and col_idx == severity_col + 1:
                v = str(value).upper()
                if v == "FAIL":
                    cell.fill = FAIL_FILL
                elif v == "WARN":
                    cell.fill = WARN_FILL
                elif v in ("CLEAN", "OK"):
                    cell.fill = CLEAN_FILL

        # Alternate row shading
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                if c.fill == PatternFill():  # Only if not already colored
                    c.fill = PatternFill("solid", fgColor="F2F2F2")

    # Auto-width (sample first 50 rows + header)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in capped[:50]:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

    # Freeze and filter
    ws.freeze_panes = "A2"
    if headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(capped) + 1}"

    return len(capped)


def _coerce(value: str) -> Any:
    """Try to convert CSV string to number for Excel."""
    if not value:
        return value
    try:
        if "." in value:
            return float(value)
        return int(value)
    except ValueError:
        return value


def find_col(headers: list, name: str) -> Optional[int]:
    """Find column index by name (0-based)."""
    try:
        return headers.index(name)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Combine CSVs from both sports
# ---------------------------------------------------------------------------

def combine_csvs(filename: str) -> tuple:
    """Combine NBA + MLB CSVs for a given filename."""
    nba_path = REVIEW_DIR / "NBA" / filename
    mlb_path = REVIEW_DIR / "MLB" / filename

    nba_h, nba_r = read_csv_rows(nba_path)
    mlb_h, mlb_r = read_csv_rows(mlb_path)

    if nba_h and mlb_h:
        # Use NBA headers as canonical (they should match)
        return nba_h, nba_r + mlb_r
    elif nba_h:
        return nba_h, nba_r
    elif mlb_h:
        return mlb_h, mlb_r
    return [], []


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

def load_summaries() -> List[Dict[str, Any]]:
    """Load summary JSONs for each sport."""
    summaries = []
    for sport in ("NBA", "MLB"):
        path = REVIEW_DIR / sport / "summary.json"
        if path.exists():
            summaries.append(json.loads(path.read_text()))
    return summaries


def build_dashboard(wb: Workbook, summaries: List[Dict]) -> None:
    ws = wb.create_sheet(title="Dashboard")

    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value="Sports Review Dashboard").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now(tz=timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")

    row = 4
    headers = ["Sport", "Status", "Normalized Rows", "Eligible Rows",
               "Total Signals", "Cross-Source", "Player Props", "Issues"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    for s in summaries:
        row += 1
        sport = s.get("sport", "?")
        status = s.get("status", "?")
        ws.cell(row=row, column=1, value=sport)
        status_cell = ws.cell(row=row, column=2, value=status)
        if status == "FAIL":
            status_cell.fill = FAIL_FILL
        elif status == "WARN":
            status_cell.fill = WARN_FILL
        elif status == "CLEAN":
            status_cell.fill = CLEAN_FILL
        status_cell.font = BOLD
        ws.cell(row=row, column=3, value=s.get("normalized_rows", 0))
        ws.cell(row=row, column=4, value=s.get("eligible_rows", 0))
        ws.cell(row=row, column=5, value=s.get("signals", 0))
        ws.cell(row=row, column=6, value=s.get("cross_source", 0))
        ws.cell(row=row, column=7, value=s.get("player_props", 0))
        issue_counts = s.get("issue_counts", {})
        ws.cell(row=row, column=8, value=sum(issue_counts.values()))

    # Issue breakdown
    row += 2
    ws.cell(row=row, column=1, value="Issue Breakdown by Sport").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="Sport").font = BOLD
    ws.cell(row=row, column=2, value="Issue Type").font = BOLD
    ws.cell(row=row, column=3, value="Count").font = BOLD
    for s in summaries:
        for issue_type, count in sorted(s.get("issue_counts", {}).items()):
            row += 1
            ws.cell(row=row, column=1, value=s.get("sport"))
            ws.cell(row=row, column=2, value=issue_type)
            ws.cell(row=row, column=3, value=count)

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# README
# ---------------------------------------------------------------------------

def build_readme(wb: Workbook) -> None:
    ws = wb.create_sheet(title="README")

    content = [
        ("Sports Review Workbook", TITLE_FONT),
        ("", None),
        ("Tab Guide — Review in This Order:", SECTION_FONT),
        ("1. Dashboard — Overall status per sport. Start here.", None),
        ("2. Data Quality Issues — FAIL items must be fixed, WARN items should be reviewed", None),
        ("3. Source Coverage — Which sources have data and which are missing", None),
        ("4. Cross-Source Signals — Multi-source consensus picks (highest value)", None),
        ("5. Normalized Picks — All individual picks from all sources", None),
        ("6. Player Props — Player prop picks with stat details", None),
        ("7. Signals Latest — Deduplicated signal-level view from ledger", None),
        ("8. Expert Records — NBA expert win/loss records from graded data", None),
        ("9. Expert by Market — NBA experts broken down by market type", None),
        ("10. MLB Experts — MLB expert records derived from signals + grades", None),
        ("11. Pattern Policy — Pattern audit with proposed keep/exclude decisions", None),
        ("12. Action Items — Auto-generated list of things needing attention", None),
        ("", None),
        ("Severity Legend:", SECTION_FONT),
        ("FAIL = Data integrity error that must be fixed (red)", None),
        ("WARN = Data quality issue to investigate (yellow)", None),
        ("CLEAN = No issues found (green)", None),
        ("", None),
        ("Key Concepts:", SECTION_FONT),
        ("sources_combo = pipe-delimited list of sources agreeing on a pick", None),
        ("wilson_lower = conservative win rate estimate (95% CI lower bound)", None),
        ("roi_eligible = has odds + graded result (needed for ROI calculation)", None),
        ("matchup_key = canonical SPORT:YYYY:MM:DD:TEAM1-TEAM2 (sorted alpha)", None),
        ("", None),
        ("This workbook is READ-ONLY. It does not modify pipeline data.", None),
    ]

    for row_idx, (text, font) in enumerate(content, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        if font:
            cell.font = font

    ws.column_dimensions["A"].width = 80


# ---------------------------------------------------------------------------
# Expert Records
# ---------------------------------------------------------------------------

def _wilson_lower(wins: int, n: int, z: float = 1.96) -> float:
    if n == 0:
        return 0.0
    import math
    p = wins / n
    denom = 1 + z * z / n
    center = p + z * z / (2 * n)
    spread = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n))
    return (center - spread) / denom


def load_nba_expert_records() -> tuple:
    """Load NBA expert records from by_expert_record.json."""
    path = REPO_ROOT / "data" / "reports" / "by_expert_record.json"
    if not path.exists():
        return [], []
    data = json.loads(path.read_text())
    rows_raw = data.get("rows", [])
    headers = ["sport", "expert", "n", "wins", "losses", "pushes", "win_pct", "wilson_lower", "net_units", "roi"]
    rows = []
    for r in sorted(rows_raw, key=lambda x: -x.get("n", 0)):
        denom = r.get("wins", 0) + r.get("losses", 0)
        rows.append([
            "NBA",
            r.get("expert", ""),
            str(r.get("n", 0)),
            str(r.get("wins", 0)),
            str(r.get("losses", 0)),
            str(r.get("pushes", 0)),
            f"{r.get('win_pct', 0):.4f}",
            f"{r.get('wilson_lower', 0):.4f}",
            f"{r.get('net_units', 0):.2f}",
            f"{r.get('roi', 0):.4f}",
        ])
    return headers, rows


def load_nba_expert_records_market() -> tuple:
    """Load NBA expert records by market from by_expert_record_market.json."""
    path = REPO_ROOT / "data" / "reports" / "by_expert_record_market.json"
    if not path.exists():
        return [], []
    data = json.loads(path.read_text())
    rows_raw = data.get("rows", [])
    headers = ["sport", "expert", "market_type", "n", "wins", "losses", "pushes", "win_pct", "wilson_lower", "net_units", "roi"]
    rows = []
    for r in sorted(rows_raw, key=lambda x: (-x.get("n", 0))):
        rows.append([
            "NBA",
            r.get("expert", ""),
            r.get("market_type", ""),
            str(r.get("n", 0)),
            str(r.get("wins", 0)),
            str(r.get("losses", 0)),
            str(r.get("pushes", 0)),
            f"{r.get('win_pct', 0):.4f}",
            f"{r.get('wilson_lower', 0):.4f}",
            f"{r.get('net_units', 0):.2f}",
            f"{r.get('roi', 0):.4f}",
        ])
    return headers, rows


def derive_mlb_expert_records() -> tuple:
    """Derive MLB expert records from signals + grades (no pre-built report)."""
    from collections import defaultdict

    signals_path = REPO_ROOT / "data" / "ledger" / "mlb" / "signals_latest.jsonl"
    grades_path = REPO_ROOT / "data" / "ledger" / "mlb" / "grades_latest.jsonl"
    if not signals_path.exists() or not grades_path.exists():
        return [], []

    grades = {}
    with grades_path.open() as f:
        for line in f:
            line = line.strip()
            if line:
                g = json.loads(line)
                grades[g.get("signal_id")] = g

    expert_data: Dict[str, Dict] = defaultdict(lambda: {"n": 0, "wins": 0, "losses": 0, "pushes": 0, "units": 0.0, "units_count": 0})
    with signals_path.open() as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            sig = json.loads(line)
            sid = sig.get("signal_id")
            grade = grades.get(sid)
            if not grade or grade.get("status") not in ("WIN", "LOSS", "PUSH"):
                continue
            result = grade["status"]
            units = grade.get("units")

            for sup in (sig.get("supports") or []):
                expert = sup.get("expert_name") or sup.get("source_id", "unknown")
                rec = expert_data[expert]
                rec["n"] += 1
                if result == "WIN":
                    rec["wins"] += 1
                elif result == "LOSS":
                    rec["losses"] += 1
                elif result == "PUSH":
                    rec["pushes"] += 1
                if units is not None:
                    rec["units"] += units
                    rec["units_count"] += 1

    headers = ["sport", "expert", "n", "wins", "losses", "pushes", "win_pct", "wilson_lower", "net_units", "roi"]
    rows = []
    for expert, rec in sorted(expert_data.items(), key=lambda x: -x[1]["n"]):
        denom = rec["wins"] + rec["losses"]
        wr = rec["wins"] / denom if denom > 0 else 0.0
        wl = _wilson_lower(rec["wins"], denom)
        roi = rec["units"] / rec["units_count"] if rec["units_count"] > 0 else 0.0
        rows.append([
            "MLB",
            expert,
            str(rec["n"]),
            str(rec["wins"]),
            str(rec["losses"]),
            str(rec["pushes"]),
            f"{wr:.4f}",
            f"{wl:.4f}",
            f"{rec['units']:.2f}",
            f"{roi:.4f}",
        ])
    return headers, rows


# ---------------------------------------------------------------------------
# Action Items
# ---------------------------------------------------------------------------

def build_action_items(wb: Workbook, summaries: List[Dict], quality_headers: list, quality_rows: list) -> int:
    ws = wb.create_sheet(title="Action Items")

    ws.cell(row=1, column=1, value="Action Items — What Needs Your Review").font = TITLE_FONT

    items = []

    # From summaries
    for s in summaries:
        sport = s.get("sport", "?")
        status = s.get("status", "?")
        if status in ("WARN", "FAIL"):
            items.append((sport, status, f"{sport} overall status is {status}",
                          "Review Dashboard and Data Quality Issues tabs"))

        for issue_type, count in s.get("issue_counts", {}).items():
            sev = "WARN"
            if "FAIL" in issue_type or "wrong" in issue_type or "noncanonical" in issue_type:
                sev = "FAIL"
            items.append((sport, sev, f"{issue_type} ({count} occurrences)",
                          f"Filter Data Quality Issues tab by sport={sport}, issue_type={issue_type}"))

    # Pattern audit items
    policy_path = AUDIT_DIR / "pattern_policy_candidates.csv"
    if policy_path.exists():
        ph, pr = read_csv_rows(policy_path)
        policy_col = find_col(ph, "proposed_policy")
        sport_col = find_col(ph, "sport")
        combo_col = find_col(ph, "source_combo")
        market_col = find_col(ph, "market_type")
        wilson_col = find_col(ph, "wilson_lower")
        if policy_col is not None:
            exclusions = [r for r in pr if r[policy_col] == "exclude_from_auto_a"]
            if exclusions:
                items.append(("ALL", "WARN", f"{len(exclusions)} patterns proposed for A-tier exclusion",
                              "Review Pattern Policy tab, filter on proposed_policy=exclude_from_auto_a"))

    # Write items
    row = 3
    action_headers = ["Sport", "Severity", "Issue", "Suggested Action"]
    for col, h in enumerate(action_headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    for sport, sev, issue, action in items:
        row += 1
        ws.cell(row=row, column=1, value=sport)
        sev_cell = ws.cell(row=row, column=2, value=sev)
        if sev == "FAIL":
            sev_cell.fill = FAIL_FILL
        elif sev == "WARN":
            sev_cell.fill = WARN_FILL
        ws.cell(row=row, column=3, value=issue)
        ws.cell(row=row, column=4, value=action)

    for col, w in enumerate([10, 10, 50, 60], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"

    return len(items)


# ---------------------------------------------------------------------------
# Validation checks
# ---------------------------------------------------------------------------

def run_checks(all_picks_headers: list, all_picks_rows: list,
               quality_headers: list, quality_rows: list) -> List[str]:
    """Run data correctness checks, return list of findings."""
    findings = []

    sport_col = find_col(all_picks_headers, "sport")
    sel_col = find_col(all_picks_headers, "selection")
    pk_col = find_col(all_picks_headers, "player_key")
    mk_col = find_col(all_picks_headers, "matchup_key")
    ek_col = find_col(all_picks_headers, "event_key")
    line_col = find_col(all_picks_headers, "line")
    date_col = find_col(all_picks_headers, "date")
    src_col = find_col(all_picks_headers, "source")

    nba_with_mlb = 0
    mlb_with_nba = 0
    null_sel = 0
    null_line = 0
    missing_pk = 0
    missing_mk = 0
    latest_by_sport_source: Dict[str, str] = {}

    for row in all_picks_rows:
        sport = row[sport_col] if sport_col is not None and sport_col < len(row) else ""
        sel = row[sel_col] if sel_col is not None and sel_col < len(row) else ""
        pk = row[pk_col] if pk_col is not None and pk_col < len(row) else ""
        mk = row[mk_col] if mk_col is not None and mk_col < len(row) else ""
        line = row[line_col] if line_col is not None and line_col < len(row) else ""
        dt = row[date_col] if date_col is not None and date_col < len(row) else ""
        src = row[src_col] if src_col is not None and src_col < len(row) else ""

        if sport == "NBA" and ("MLB:" in sel or "MLB:" in pk):
            nba_with_mlb += 1
        if sport == "MLB" and ("NBA:" in sel or "NBA:" in pk):
            mlb_with_nba += 1
        if not sel:
            null_sel += 1
        if not line:
            null_line += 1
        if not pk and "prop" in str(row):
            missing_pk += 1
        if not mk:
            missing_mk += 1

        key = f"{sport}|{src}"
        if dt and (key not in latest_by_sport_source or dt > latest_by_sport_source[key]):
            latest_by_sport_source[key] = dt

    findings.append(f"MLB rows with NBA: prefix: {mlb_with_nba}")
    findings.append(f"NBA rows with MLB: prefix: {nba_with_mlb}")
    findings.append(f"Null selections: {null_sel}")
    findings.append(f"Null lines: {null_line}")
    findings.append(f"Missing matchup_key: {missing_mk}")
    findings.append(f"Missing player_key (in prop-like rows): {missing_pk}")
    findings.append(f"Latest date by sport|source:")
    for key in sorted(latest_by_sport_source):
        findings.append(f"  {key}: {latest_by_sport_source[key]}")

    return findings


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("  EXPORT REVIEW EXCEL WORKBOOK")
    print("=" * 60)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    tab_counts = {}

    # 1. README
    build_readme(wb)
    tab_counts["README"] = 0

    # 2. Dashboard — build from CSVs since summary.json may not exist
    summaries = load_summaries()
    if not summaries:
        # Derive summaries from CSV data
        for sport in ("NBA", "MLB"):
            s_dir = REVIEW_DIR / sport
            norm_h, norm_r = read_csv_rows(s_dir / "normalized_picks.csv")
            sig_h, sig_r = read_csv_rows(s_dir / "signals_latest.csv")
            xs_h, xs_r = read_csv_rows(s_dir / "cross_source_signals.csv")
            prop_h, prop_r = read_csv_rows(s_dir / "player_props.csv")
            dq_h, dq_r = read_csv_rows(s_dir / "data_quality_issues.csv")

            elig_col = find_col(norm_h, "eligible_for_consensus")
            eligible = sum(1 for r in norm_r if elig_col is not None and elig_col < len(r)
                           and str(r[elig_col]).lower() in ("true", "1", "yes"))

            sev_col = find_col(dq_h, "severity")
            issue_type_col = find_col(dq_h, "issue_type")
            ic: Dict[str, int] = {}
            has_fail = False
            for r in dq_r:
                it = r[issue_type_col] if issue_type_col is not None and issue_type_col < len(r) else "?"
                ic[it] = ic.get(it, 0) + 1
                if sev_col is not None and sev_col < len(r) and r[sev_col] == "FAIL":
                    has_fail = True
            status = "FAIL" if has_fail else ("WARN" if ic else "CLEAN")

            summaries.append({
                "sport": sport,
                "status": status,
                "normalized_rows": len(norm_r),
                "eligible_rows": eligible,
                "signals": len(sig_r),
                "cross_source": len(xs_r),
                "player_props": len(prop_r),
                "issue_counts": ic,
            })
    build_dashboard(wb, summaries)
    tab_counts["Dashboard"] = len(summaries)

    # 3. Source Coverage
    h, r = combine_csvs("source_coverage.csv")
    n = add_data_sheet(wb, "Source Coverage", h, r)
    tab_counts["Source Coverage"] = n

    # 4. All Normalized Picks
    h, r = combine_csvs("normalized_picks.csv")
    n = add_data_sheet(wb, "Normalized Picks", h, r)
    tab_counts["Normalized Picks"] = n
    all_picks_h, all_picks_r = h, r

    # 5. Signals Latest
    h, r = combine_csvs("signals_latest.csv")
    n = add_data_sheet(wb, "Signals Latest", h, r)
    tab_counts["Signals Latest"] = n

    # 6. Cross-Source Signals
    h, r = combine_csvs("cross_source_signals.csv")
    n = add_data_sheet(wb, "Cross-Source", h, r)
    tab_counts["Cross-Source"] = n

    # 7. Player Props
    h, r = combine_csvs("player_props.csv")
    n = add_data_sheet(wb, "Player Props", h, r)
    tab_counts["Player Props"] = n

    # 8. Data Quality Issues
    h, r = combine_csvs("data_quality_issues.csv")
    sev_col = find_col(h, "severity")
    n = add_data_sheet(wb, "Data Quality Issues", h, r, severity_col=sev_col)
    tab_counts["Data Quality Issues"] = n
    quality_h, quality_r = h, r

    # 9. Expert Records — NBA
    exp_h, exp_r = load_nba_expert_records()
    n = add_data_sheet(wb, "Expert Records", exp_h, exp_r)
    tab_counts["Expert Records"] = n

    # 10. Expert Records by Market — NBA
    expm_h, expm_r = load_nba_expert_records_market()
    n = add_data_sheet(wb, "Expert by Market", expm_h, expm_r)
    tab_counts["Expert by Market"] = n

    # 11. MLB Expert Records — derived from signals + grades
    mlb_exp_h, mlb_exp_r = derive_mlb_expert_records()
    # Combine NBA + MLB expert records into one "All Experts" tab
    if exp_h and mlb_exp_h:
        all_exp_rows = exp_r + mlb_exp_r
        n = add_data_sheet(wb, "All Experts", exp_h, all_exp_rows)
        tab_counts["All Experts"] = n
    elif mlb_exp_h:
        n = add_data_sheet(wb, "All Experts", mlb_exp_h, mlb_exp_r)
        tab_counts["All Experts"] = n

    # 12. MLB Experts standalone
    n = add_data_sheet(wb, "MLB Experts", mlb_exp_h, mlb_exp_r)
    tab_counts["MLB Experts"] = n

    # 13. NBA Picks
    nba_h, nba_r = read_csv_rows(REVIEW_DIR / "NBA" / "normalized_picks.csv")
    n = add_data_sheet(wb, "NBA Picks", nba_h, nba_r)
    tab_counts["NBA Picks"] = n

    # 14. MLB Picks
    mlb_h, mlb_r = read_csv_rows(REVIEW_DIR / "MLB" / "normalized_picks.csv")
    n = add_data_sheet(wb, "MLB Picks", mlb_h, mlb_r)
    tab_counts["MLB Picks"] = n

    # 15. Pattern Policy (if exists)
    policy_path = AUDIT_DIR / "pattern_policy_candidates.csv"
    if policy_path.exists():
        h, r = read_csv_rows(policy_path)
        policy_col = find_col(h, "proposed_policy")
        n = add_data_sheet(wb, "Pattern Policy", h, r)
        tab_counts["Pattern Policy"] = n
    else:
        tab_counts["Pattern Policy"] = 0

    # Pattern performance
    perf_path = AUDIT_DIR / "pattern_performance_by_sport.csv"
    if perf_path.exists():
        h, r = read_csv_rows(perf_path)
        n = add_data_sheet(wb, "Pattern Performance", h, r)
        tab_counts["Pattern Performance"] = n

    # 12. Action Items (use latest summaries, derived or loaded)
    n = build_action_items(wb, summaries, quality_h, quality_r)
    tab_counts["Action Items"] = n

    # Run validation checks
    print("\n  Data Correctness Checks:")
    findings = run_checks(all_picks_h, all_picks_r, quality_h, quality_r)
    for f in findings:
        print(f"    {f}")

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))

    print(f"\n  Workbook: {OUTPUT_PATH}")
    print(f"\n  Tabs created:")
    for tab, count in tab_counts.items():
        print(f"    {tab:25s} {count:>6} rows")

    total_rows = sum(tab_counts.values())
    print(f"\n  Total data rows: {total_rows}")

    # Warnings summary
    warn_count = sum(1 for r in quality_r if any("WARN" in c for c in r))
    fail_count = sum(1 for r in quality_r if any("FAIL" in c for c in r))
    print(f"\n  Remaining warnings: {warn_count} WARN, {fail_count} FAIL")

    print(f"\n  Review priority:")
    print(f"    1. Dashboard tab — check sport status")
    print(f"    2. Data Quality Issues tab — investigate FAIL then WARN")
    print(f"    3. Cross-Source tab — verify multi-source signals look correct")
    print(f"    4. Pattern Policy tab — review proposed exclusions")
    print(f"    5. Action Items tab — checklist of all items needing attention")


if __name__ == "__main__":
    main()
